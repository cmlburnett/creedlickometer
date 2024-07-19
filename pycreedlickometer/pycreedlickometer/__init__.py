import csv
import datetime
import functools
import itertools
import os

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from matplotlib import pyplot
import matplotlib

import pandas as pd
import numpy as np

__all__ = ['StatBot', 'CreedLickometer', 'VolumeData']

class StatBot:
	"""
	Do a bunch of basic stats on data
	"""
	def __init__(self, dat):
		self.Data = list(dat)
		self.Length = len(self.Data)
		self.Sum = sum(self.Data)

		if self.Length == 0:
			self.Minimum = None
			self.Maximum = None
			self.MinMax = None
			self.Span = None

			self.Mean = None
			self.Quartile25 = None
			self.Median = None
			self.Quartile75 = None
			self.IQR = None
		else:
			self.Minimum = min(self.Data)
			self.Maximum = max(self.Data)
			self.MinMax = (self.Minimum, self.Maximum)
			self.Span = self.Maximum - self.Minimum
			self.Mean = self.Sum / self.Length

			q = np.quantile(self.Data, [0.25,0.5,0.75])
			self.Quartile25 = q[0]
			self.Median = q[1]
			self.Quartile75 = q[2]
			self.IQR = q[2]-q[0]

class TimeData:
	"""
	Manages light & dark cycles.
	This ensures the full 24 hours is covered with a phase of the cycle.
	"""
	def __init__(self):
		self.cycles = []
		self.IsComplete = False
		self.IsProcessed = False

	def AddLightPhase(self, start, end):
		"""
		Add a light phase from @start to @end.
		If @end is before start then it is considered to have spanned midnight and wraps around.
		"""
		if start == end:
			raise ValueError("Start and end are the same time: %s" % start)

		if start < end:
			self.cycles.append( (start,end, 'light') )
		else:
			self.cycles.append( (datetime.time.min, end, 'light') )
			self.cycles.append( (start, datetime.time.max, 'light') )

		self._Process()

	def AddDarkPhase(self, start, end):
		"""
		Add a dark phase from @start to @end.
		If @end is before start then it is considered to have spanned midnight and wraps around.
		"""
		if start == end:
			raise ValueError("Start and end are the same time: %s" % start)

		if start < end:
			self.cycles.append( (start,end, 'dark') )
		else:
			self.cycles.append( (datetime.time.min, end, 'dark') )
			self.cycles.append( (start, datetime.time.max, 'dark') )

		self._Process()

	def _Process(self):
		"""
		Internal process to sort and check if complete
		"""
		self.IsComplete = False
		self.IsProcessed = False

		if not len(self.cycles):
			return

		self.cycles.sort()

		if self.cycles[0][0] != datetime.time.min:
			return
		if self.cycles[-1][1] != datetime.time.max:
			return

		# Find if any gaps between parts
		for idx in range(1,len(self.cycles)):
			if self.cycles[idx-1][1] != self.cycles[idx][0]:
				break

		# If it made it through then it's complete
		self.IsComplete = True

	def Process(self):
		"""
		Called by the user to formally check midnight and complete clock coverage with Exceptions thrown if not.
		"""

		self._Process()

		if self.cycles[0][0] != datetime.time.min:
			raise ValueError("TimeData does not start at midnight")
		if self.cycles[-1][1] != datetime.time.max:
			raise ValueError("TimeData does not end at midnight")

		if not self.IsComplete:
			raise ValueError("TimeData has a gap")

		self.IsProcessed = True

	def GetTime(self, dt):
		if isinstance(dt, datetime.datetime):
			dt = dt.time()
		elif isinstance(dt, datetime.time):
			pass
		else:
			raise TypeError("Expect datetime or time object, got %d instead" % str(dt))

		for start,end,phase in self.cycles:
			if start <= dt and dt < end:
				return phase
		else:
			raise ValueError("Found time %s not in time data, which shouldn't happen" % str(dt))

class VolumeData:
	"""
	Manages volume data.
	Data can come in the form of measured data or fill data.
	Fill data should be AFTER meausred data (gotta have second point to determine velocity).

	Accepted for measured and fill data to be at the same time, so a 1 microsecond delta is added to the measured time
	 so that fill comes second when sorting.
	"""
	def __init__(self):
		# Tuples of (datetime, code, device ID, left, right)
		# code is just 'm' or 'f' for measure or fill
		self.measured = []
		self.fill = []
		self.process = []

	def AddMeasurement(self, dt, device, left, right):
		"""
		Add a measurement value at time @dt for device @device for the left and right tubes.
		Left and right to be assumed to be milliliters.
		"""
		self.measured.append( (dt, 'm', device, left, right) )
		self._Process()

	def AddFill(self, dt, device, left, right):
		"""
		Add a fill value at time @dt for device @device for the left and right tubes.
		Left and right to be assumed to be milliliters.
		"""
		self.fill.append( (dt + datetime.timedelta(microseconds=1), 'f', device, left, right) )
		self._Process()

	def _Process(self):
		"""
		Internal function called with every addition of data.
		Process the entries and collapse measured/fill data.
		"""

		self.measured.sort(key=lambda _:_[0])
		self.fill.sort(key=lambda _:_[0])

		self.process.clear()
		self.process += self.measured
		self.process += self.fill
		self.process.sort()

	def GetVolume(self, dt, device):
		"""
		Get volume information on device @device at the time @dt.
		Returns a dictionary of pre, post, and delta data.
		All three return values are 3-tuples of (date, left, right) data.
		"""

		if not len(self.measured):
			raise ValueError("No volume data")

		# Exclude all other devices
		filtered = [_ for _ in self.process if _[2] == device]

		# Find bound entires of the time
		for idx,entry in enumerate(filtered):
			if dt < entry[0]:
				if idx == 0:
					raise ValueError("Time requested (%s) is before all available data, cannot give a volume" % dt)
				else:
					for lidx in range(1,idx):
						pre_l = filtered[idx-lidx]
						if pre_l[3] is not None:
							break
					else:
						raise ValueError("Unable to find a prior left value starting at index %d" % idx)

					for ridx in range(1,idx):
						pre_r = filtered[idx-ridx]
						if pre_r[4] is not None:
							break
					else:
						raise ValueError("Unable to find a prior right value starting at index %d" % idx)

					post = entry

					return {
						# This should be a unique index per device but is meaningless to compare before/after adding data
						'index': idx,
						# Pre and post (datetime, left, right)
						'pre': (pre_l[0], pre_l[3], pre_r[4]),
						'post': (post[0], post[3], post[4]),
						# Change in time and volumes of this time span
						'delta': (post[0]-pre_l[0], pre_l[3]-post[3], pre_r[4]-post[4])
					}
			else:
				# Haven't bounded yet, loop again
				pass
		else:
			# After last entry
			raise ValueError("Time requested (%s) is after all available data, cannot give a volume" % dt)

class CreedLickometer:
	def __init__(self, fname):
		self.Filename = fname

		# No volume data by default
		self.VolumeData = None

		# No time cycle data by default
		self.TimeData = None

		# Integer of the programmed device ID
		self.DeviceID = None

		# Tuples of (minimum,maximum) values for datetime and milliseconds values
		self.Spandt = None
		self.Spanms = None

		# Raw data of (datetime, milliseconds integer, Beam open, delta)
		# where Beam open is True == beam opened, False == beam closed
		# and delta == change from prior state (either a bout time or interbout time)
		self.Lefts = None
		self.Rights = None

		# List of bout deltas
		self.LeftBouts = None
		self.RightBouts = None

		# List of intebout deltas
		self.LeftInterbouts = None
		self.RightInterbouts = None

		# Dictionary mapping datetime by minute to bout deltas
		self.LeftVsTime = None
		self.RightVsTime = None

		# Cumulative data
		self.LeftCumulative = None
		self.RightCumulative = None

		# Cumulative volume data (for each measured/fill phase)
		self.LeftCumulativeVolume = None
		self.RightCumulativeVolume = None

		# Cumulative volume data (total)
		self.LeftCumulativeTotalVolume = None
		self.RightCumulativeTotalVolume = None

		self.IsMerged = False
		self.IsLoaded = False
		self.IsProcessed = False

	def AddTimeData(self, tz):
		"""
		Time data provides light/dark cycle information.
		"""
		self.TimeData = tz

		# Force that it's checked
		if tz.IsProcessed:
			tz.Process()

	def AddVolumeData(self, volume):
		"""
		Volume data provides fill and measured data for some plots.
		If no volume data is provided, then those plots aren't generated.
		"""
		self.VolumeData = volume

		# Reprocess if volume data is set
		if self.IsProcessed:
			self.Process()

	def __repr__(self):
		return "<%s device=%s file=%s>" % (self.__class__.__name__, self.DeviceID, self.Filename)

	def Save(self, fname=None):
		"""
		Save the data to CSV file.
		Optional file name to save to (using this saves a copy and leaves self.Filename alone).
		Battery voltage data is discarded, so update original data files at your own peril.
		"""

		rows = []
		if not len(self.Rights) and not len(self.Lefts):
			# No data to write
			pass

		# No rights, just lefts
		elif not len(self.Rights):
			for row in self.Lefts:
				l = int(not row[2])
				r = 1
				rows.append( [row[0], row[1], self.DeviceID, l, r, 0.0] )

		# No lefts, only rights
		elif not len(self.Lefts):
			for row in self.Rights:
				l = 1
				r = int(not row[2])
				rows.append( [row[0], row[1], self.DeviceID, l, r, 0.0] )

		else:
			# Combine left and rights
			combo = [[dt,ms,int(not beam),None] for dt,ms,beam,delta in self.Lefts]
			combo +=[[dt,ms,None,int(not beam)] for dt,ms,beam,delta in self.Rights]
			# Sort by milliseconds
			combo.sort(key=lambda _:_[1])

			# Keep track of last left/right values to copy on rows in which it doesn't change
			l = 1
			r = 1
			for row in combo:
				# Update left/right value depending on if it was set or not
				if row[2] is not None:
					l = row[2]
				if row[3] is not None:
					r = row[3]

				rows.append((
					row[0].strftime("%Y-%m-%d %H:%M:%S"),
					row[1],
					self.DeviceID,
					l,
					r,
					0.0
				))

		header = ['YYYY-MM-DD hh:mm:ss', 'Millseconds', 'Device', 'LeftState', 'RightState', 'BatteryVoltage']

		# Overwrite file
		if fname is None:
			fname = self.Filename

		with open(fname, 'w') as f:
			w = csv.writer(f)
			w.writerow(header)

			for row in rows:
				w.writerow(row)

	def Load(self):
		"""
		Load a CSV data file without processing
		"""

		lefts = []
		rights = []

		with open(self.Filename, 'r') as f:
			r = csv.reader(f)
			for row in r:
				# Header row, disregard
				if row[0].startswith("YYYY"):
					continue

				try:
					dt = datetime.datetime.strptime(row[0], '%m/%d/%Y %H:%M')
				except ValueError:
					dt = datetime.datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')

				ms = int(row[1])
				self.DeviceID = int(row[2])
				left = int(row[3])
				right = int(row[4])

				if not len(lefts) and not len(rights):
					if left == 1 and right == 1:
						lefts.append( (dt,ms,False,None) )
						rights.append( (dt,ms,False,None) )
					else:
						# Need to find a (1,1) row indicating neither are blocked
						continue
				else:
					# Scenarios:
					#  1) Beam open (False) and still is open (1)
					#  2) Beam open (False) and is closed (0) [MOUSE STARTS DRINKING]
					#  3) Beam closed (True) and is open (1) [MOUSE STOPS DRINKING]
					#  4) Beam closed (True) and still is closed (0)

					# (2)
					if lefts[-1][2] == False and left == 0:
						delta = ms - lefts[-1][1]
						lefts.append( (dt,ms,True, delta) )

					# (3)
					elif lefts[-1][2] == True and left == 1:
						delta = ms - lefts[-1][1]
						lefts.append( (dt,ms,False, delta) )
					else:
						pass

					# (2)
					if rights[-1][2] == False and right == 0:
						delta = ms - rights[-1][1]
						rights.append( (dt,ms,True, delta) )

					# (3)
					elif rights[-1][2] == True and right == 1:
						delta = ms - rights[-1][1]
						rights.append( (dt,ms,False, delta) )
					else:
						pass

		self.Lefts = lefts
		self.Rights = rights
		self.IsLoaded = True

	def TrimBefore(self, truncate_dt):
		"""
		Trim all data before datetime @truncate_dt.
		"""

		if not self.IsLoaded:
			self.Load()

		# Pull out trimmed entries
		lefts = []
		rights = []

		for dt,ms,beam,delta in self.Lefts:
			if dt >= truncate_dt:
				lefts.append( (dt,ms,beam,delta) )
		for dt,ms,beam,delta in self.Rights:
			if dt >= truncate_dt:
				rights.append( (dt,ms,beam,delta) )

		# Edge case of the truncation date being in the middle of a bout
		# If it starts in the beam broken state ([2] == True) then exclude it
		if lefts[0][2]:
			del lefts[0]
		if rights[0][2]:
			del rights[0]

		# Create new container for the data, assign it, and pretend it's loaded
		o = CreedLickometer(None)
		o.DeviceID = self.DeviceID
		o.VolumeData = self.VolumeData
		o.TimeData = self.TimeData
		o.Lefts = lefts
		o.Rights = rights
		o.IsLoaded = True
		o.IsMerged = True

		return o

	def TrimAfter(self, truncate_dt):
		"""
		Trim all data after datetime @truncate_dt.
		"""

		if not self.IsLoaded:
			self.Load()

		# Pull out trimmed entries
		lefts = []
		rights = []

		for dt,ms,beam,delta in self.Lefts:
			if dt <= truncate_dt:
				lefts.append( (dt,ms,beam,delta) )
		for dt,ms,beam,delta in self.Rights:
			if dt <= truncate_dt:
				rights.append( (dt,ms,beam,delta) )

		# Edge case of the truncation date being in the middle of a bout
		# If it starts in the beam broken state ([2] == True) then exclude it
		if lefts[-1][2]:
			lefts.pop()
		if rights[-1][2]:
			rights.pop()

		# Create new container for the data, assign it, and pretend it's loaded
		o = CreedLickometer(None)
		o.DeviceID = self.DeviceID
		o.VolumeData = self.VolumeData
		o.TimeData = self.TimeData
		o.Lefts = lefts
		o.Rights = rights
		o.IsLoaded = True
		o.IsMerged = True

		return o

	@staticmethod
	def Merge(a, b):
		"""
		Merge two data files @a and @b and return a new CreedLickometer instance
		"""

		# Ensure files are loaded
		if not a.IsLoaded:
			a.Load()
		if not b.IsLoaded:
			b.Load()

		# Ensure files are processed (I'm lazy and don't want to calculate Spandt myself)
		if not a.IsProcessed:
			a.Process()
		if not b.IsProcessed:
			b.Process()

		# Why would you merge files from different devices other than by accident?
		if a.DeviceID != b.DeviceID:
			raise ValueError("Merging files from two different devices (%d and %d)" % (a.DeviceID, b.DeviceID))

		# Ensure same volume data for both data sets
		if a.VolumeData is None and b.VolumeData is None:
			pass
		elif a.VolumeData is not None and b.VolumeData is not None:
			if id(a.VolumeData) != id(b.VolumeData):
				raise ValueError("Merging two sets of data with different volume data")
			else:
				# Same object, so move on
				pass
		else:
			raise ValueError("Merging two sets of data but one does not have volume data and the other does")

		# Ensure same volume data for both data sets
		if a.TimeData is None and b.TimeData is None:
			pass
		elif a.TimeData is not None and b.TimeData is not None:
			if id(a.TimeData) != id(b.TimeData):
				raise ValueError("Merging two sets of data with different time data")
			else:
				# Same object, so move on
				pass
		else:
			raise ValueError("Merging two sets of data but one does not have time data and the other does")

		# -----------------------------------------------------------------------------------------
		# -----------------------------------------------------------------------------------------

		# No data, so no need to merge
		if a.Spandt[0] is None:
			return b
		elif b.Spandt[0] is None:
			return a

		# Flip file order if wrong
		if b.Spandt[0] >= a.Spandt[1]:
			# a <= b is correctly ordered as expected
			pass
		elif a.Spandt[0] > b.Spandt[1]:
			# b < a so swap them
			a,b = b,a
		else:
			raise ValueError("Unrecognized ordering of files: a=%s, b=%s" % (a.Spandt, b.Spandt))

		# Time gap between files
		gap = b.Spandt[0] - a.Spandt[1]

		# Not a halting error, just disregard data we can't finish processing
		if a.Lefts[-1][2]:
			print("First file (%s) left ends with beam broken, popping this off" % a.Filename)
			a.Lefts.pop()
		if b.Lefts[0][2]:
			print("Second file (%s) left starts with a beam broken, dequeueing it" % b.Filename)
			del b.Lefts[0]

		if a.Rights[-1][2]:
			print("First file (%s) right ends with beam broken, popping this off" % a.Filename)
			a.Rights.pop()
		if b.Rights[0][2]:
			print("Second file (%s) right starts with a beam broken, dequeueing it" % b.Filename)
			del b.Rights[0]

		# -----------------------------------------------------------------------------------------
		# -----------------------------------------------------------------------------------------

		# Expected gap given gap in seconds
		gapms = int(gap.total_seconds() * 1000 + 1000)
		startms = gapms + a.Spanms[1]
		deltams = startms - b.Spanms[0]

		# Aggregate lefts & rights together
		lefts = list(a.Lefts)
		rights = list(a.Rights)

		# Merge left & right data and adjust @b's milliseconds to account for the gap
		# This ensures that millseconds is increasing across the merged data sets
		for dt,ms,beam,delta in b.Lefts:
			lefts.append( (dt,ms+deltams,beam,delta) )
		for dt,ms,beam,delta in b.Rights:
			rights.append( (dt,ms+deltams,beam,delta) )

		# Create new container for the data, assign it, and pretend it's loaded
		o = CreedLickometer(None)
		o.DeviceID = a.DeviceID
		o.VolumeData = a.VolumeData
		o.TimeData = a.TimeData
		o.Lefts = lefts
		o.Rights = rights
		o.IsLoaded = True
		o.IsMerged = True
		o.Process()

		return o

	def Process(self):
		"""
		Process the raw data in Lefts & Rights into the various parts and data.
		"""

		# Clear everything
		self.LeftBouts = []
		self.RightBouts = []

		self.LeftInterbouts = []
		self.RightInterbouts = []

		self.LeftVsTime = {}
		self.RightVsTime = {}

		self.LeftCumulative = []
		self.RightCumulative = []

		self.LeftCumulativeVolume = []
		self.RightCumulativeVolume = []

		self.LeftCumulativeTotalVolume = []
		self.RightCumulativeTotalVolume = []

		print("-"*140)
		print(self.Filename)
		print("-"*140)

		def topandas(lr_idx, entries, volume_pdf):
			bouts = []
			ld_phase_last = None
			ld_phase_idx = 0
			for idx,row in enumerate(entries):
				dt,ms,beam,delta = row
				if not beam:
					priorrow = entries[idx-1]
					# Shouldn't compare
					if not priorrow[2]:
						continue

					try:
						voldat = self.VolumeData.GetVolume(dt, self.DeviceID)
					except ValueError as e:
						print([dt,e])
						continue
					lightdat = self.TimeData.GetTime(dt)

					# Calculate a running phase index of light/datk so that it can be grouped
					if ld_phase_last is None:
						ld_phase_last = lightdat
					if lightdat != ld_phase_last:
						ld_phase_last = lightdat
						ld_phase_idx += 1

					if voldat['index'] not in volume_pdf:
						volume_pdf[ voldat['index'] ] = voldat['delta'][1+lr_idx]

					bouts.append({
						'start_dt': priorrow[0],
						'start_ms': priorrow[1],
						'end_dt': dt,
						'end_ms': ms,
						'delta': ms - priorrow[1],
						'volume': voldat['delta'][1+lr_idx],
						'volume_index': voldat['index'],
						'light': lightdat == 'light',
						'light_idx': ld_phase_idx
					})
			return pd.DataFrame(bouts)

		# Map the volume_index to volume as a "pdf" then cumulative sum those values to make a "cdf"
		# This can be added to the cumulative_volume at each row to get cumulative_total_volume
		left_volume_pdf = {}
		left_volume_cdf = {}
		right_volume_pdf = {}
		right_volume_cdf = {}

		left = topandas(0, self.Lefts, left_volume_pdf)
		right = topandas(1, self.Rights, right_volume_pdf)

		# Calculate volume CDF's
		c = 0.0
		for k in left_volume_pdf.keys():
			left_volume_cdf[k] = c
			c += left_volume_pdf[k]

		c = 0.0
		for k in right_volume_pdf.keys():
			right_volume_cdf[k] = c
			c += right_volume_pdf[k]

		# Assume an empty file
		mindt = maxdt = minms = maxms = None

		def o(data, cdf):
			# Calculate the cumulative sum of the delta (this sums the delta for the entire series)
			data['delta_total_cdf'] = data['delta'].cumsum()
			# Best way I found to do this is to copy the volume index first, them map that column using the CDF data generated
			data['volume_cumulative_base'] = data['volume_index']
			data['volume_cumulative_base'] = data['volume_cumulative_base'].map(cdf.get)

			grp = data.groupby('volume_index')
			data['delta_cdf'] = grp['delta'].cumsum()

			# Create new columns of data ultimately to get cumulative_volume and cumulative_total_volume
			def f1(g):
				g['vol_delta_pdf'] = g['delta'] * g['volume']
				g['vol_delta_cdf'] = g['delta_cdf'] * g['volume']
				g['step_volume'] = g['vol_delta_pdf'] / g['delta'].sum()
				g['cumulative_volume'] = g['vol_delta_cdf'] / g['delta'].sum()
				g['cumulative_total_volume'] = g['cumulative_volume'] + g['volume_cumulative_base']
				return g

			# Calculate the other values and back copy into @data (have to do droplevel(0) to reduce index so data can be inserted back into the frame)
			#   https://stackoverflow.com/questions/20737811/attaching-a-calculated-column-to-an-existing-dataframe-raises-typeerror-incompa
			g = grp.apply(f1)
			data['vol_delta_pdf'] = g['vol_delta_pdf'].droplevel(0)
			data['vol_delta_cdf'] = g['vol_delta_cdf'].droplevel(0)
			data['step_volume'] = g['step_volume'].droplevel(0)
			data['cumulative_volume'] = g['cumulative_volume'].droplevel(0)
			data['cumulative_total_volume'] = g['cumulative_total_volume'].droplevel(0)

			# Split up by light phase index and create pdf function (no need for cdf of this)
			grp = data.groupby('light_idx')
			def f2(g):
				g['lightdark_phase_volume_pdf'] = g['step_volume']
				return g

			y = grp.apply(f2)
			data['lightdark_phase_volume_pdf'] = y['lightdark_phase_volume_pdf'].droplevel(0)

			# No group by light or dark phase to create cdf for each phase over the whole data range
			grp = data.groupby('light')
			data['lightdark_phase_total_volume_cdf'] = grp['lightdark_phase_volume_pdf'].cumsum()

			data['light_phase_total_volume_cdf'] = None
			data['dark_phase_total_volume_cdf'] = None

			# Using a mask of light and dark phase to copy data from lightdark_phase_total_volume_cdf
			# Eseentially just splits 'lightdark_phase_total_volume_cdf' into two columns
			mask = data['light'] == True
			data.loc[mask, 'light_phase_total_volume_cdf'] = data.loc[mask, 'lightdark_phase_total_volume_cdf']
			mask = data['light'] == False
			data.loc[mask, 'dark_phase_total_volume_cdf'] = data.loc[mask, 'lightdark_phase_total_volume_cdf']

			return data

		# Process left data
		if len(left):
			mindt = left['start_dt'].min()
			maxdt = left['end_dt'].max()
			minms = left['start_ms'].min()
			maxms = left['end_ms'].max()

			left = o(left, left_volume_cdf)

		# Process right data
		if len(right):
			if len(left):
				mindt = min(mindt, right['start_dt'].min())
				maxdt = max(maxdt, right['end_dt'].max())
				minms = min(minms, right['start_ms'].min())
				maxms = max(maxms, right['end_ms'].max())
			else:
				mindt = right['start_dt'].min()
				maxdt = right['end_dt'].max()
				minms = right['start_ms'].min()
				maxms = right['end_ms'].max()

			right = o(right, right_volume_cdf)

		if False:
			# Processing performed in two passes
			#  1) Crunch data into "slices" where for second pass.
			#  2) Crunch volume data into cumulative volume data
			#  3) Crunch light-dark cycle information

			def firstpass(entries, vstime, cumulative, bouts, interbouts):
				allslices = []
				slices = []
				allslices.append(slices)

				# Keep track of when the date changes
				previous_dt = None

				t = 0.0
				for dt,ms,beam,delta in entries:
					if delta is None: continue

					# Make sure both start at time zero
					if not len(cumulative):
						cumulative.append( (dt,t) )

					if beam == True:
						interbouts.append(delta)
						cumulative.append( (dt,t) )
					else:
						bouts.append(delta)

						t += delta
						cumulative.append( (dt,t) )
						if dt not in vstime:
							vstime[dt] = []
						vstime[dt].append(delta)

						try:
							ret = self.VolumeData.GetVolume(dt, self.DeviceID)
							if previous_dt is None:
								previous_dt = ret['pre'][0]
							if ret['pre'][0] != previous_dt:
								slices = []
								allslices.append(slices)
								previous_dt = ret['pre'][0]
							slices.append( (dt,delta,t,ret) )

						except ValueError as e:
							print(e)
							continue

				return (t,allslices)

			def secondpass(lr_idx, allslices, cumulative_volumes, cumulative_total_volumes):
				data_1 = 0.0

				t1 = 0.0
				t2 = 0.0
				for idx in range(len(allslices)):
					slices = allslices[idx]

					last_val = 0.0
					if idx != 0:
						last_val = allslices[idx-1][-1][2]

					if not len(slices):
						continue

					for dt,delta,t,voldat in slices:
						vol = voldat['delta'][lr_idx+1]

						if not len(cumulative_volumes):
							cumulative_volumes.append( (dt, 0.0) )
						if not len(cumulative_total_volumes):
							cumulative_total_volumes.append( (dt, 0.0) )

						# Because @t is cumulative, have to subtract from numerator AND denominator
						t1 = vol_cumulative = (t-last_val)*vol / (slices[-1][2]-last_val)

						t2 = vol_total_cumulative = vol_cumulative + data_1

						z = datetime.timedelta(milliseconds=delta)
						cumulative_volumes.append( (dt-z, cumulative_volumes[-1][1]) )
						cumulative_volumes.append( (dt, vol_cumulative) )

						cumulative_total_volumes.append( (dt-z, cumulative_total_volumes[-1][1]) )
						cumulative_total_volumes.append( (dt, vol_total_cumulative) )

					# Save this last one so it can be added into vol_total_cumulative for next slices block
					data_1 = vol_cumulative

				return (t1,t2)

			def thirdpass(lf_idx, allslices):
				if self.DeviceID != 1:
					return

				lights = []
				darks = []

				for idx in range(len(allslices)):
					slices = allslices[idx]

					for dt,delta,t,voldat in slices:
						ret = self.TimeData.GetTime(dt.time())
						if ret == 'light':
							lights.append( (dt,delta,t,voldat) )
						elif ret == 'dark':
							darks.append( (dt,delta,t,voldat) )
						else:
							raise ValueError("Unknown light-dark phase for time %s: %d" % (str(dt),ret))

				print(lights)
				print(darks)

			left_t,left_allslices = firstpass(self.Lefts, self.LeftVsTime, self.LeftCumulative, self.LeftBouts, self.LeftInterbouts)
			right_t,right_allslices = firstpass(self.Rights, self.RightVsTime, self.RightCumulative, self.RightBouts, self.RightInterbouts)

			left_c_vol,left_ct_vol = secondpass(0, left_allslices, self.LeftCumulativeVolume, self.LeftCumulativeTotalVolume)
			right_c_vol,right_ct_vol = secondpass(1, right_allslices, self.RightCumulativeVolume, self.RightCumulativeTotalVolume)

			thirdpass(0, left_allslices)
			#thirdpass(1, right_allslices)

			# Not the most efficient way but easy to write
			mindt = min(map(lambda _:_[0], self.Lefts + self.Rights))
			maxdt = max(map(lambda _:_[0], self.Lefts + self.Rights))
			minms = min(map(lambda _:_[1], self.Lefts + self.Rights))
			maxms = max(map(lambda _:_[1], self.Lefts + self.Rights))

		# Set the spans of the time data
		self.Spandt = (mindt, maxdt)
		self.Spanms = (minms, maxms)

		if False:
			# Ensure start and end are the same
			self.LeftCumulative.insert(0, (mindt, 0.0) )
			self.LeftCumulative.append( (maxdt, left_t) )
			self.RightCumulative.insert(0, (mindt, 0.0) )
			self.RightCumulative.append( (maxdt, right_t) )

			self.LeftCumulativeVolume.insert(0, (mindt, 0.0) )
			self.LeftCumulativeVolume.append( (maxdt, left_c_vol) )
			self.RightCumulativeVolume.insert(0, (mindt, 0.0) )
			self.RightCumulativeVolume.append( (maxdt, right_c_vol) )

			self.LeftCumulativeTotalVolume.insert(0, (mindt, 0.0) )
			self.LeftCumulativeTotalVolume.append( (maxdt, left_ct_vol) )
			self.RightCumulativeTotalVolume.insert(0, (mindt, 0.0) )
			self.RightCumulativeTotalVolume.append( (maxdt, right_ct_vol) )

			# Sort the data
			self.LeftBouts.sort()
			self.LeftInterbouts.sort()
			self.RightBouts.sort()
			self.RightInterbouts.sort()

			# Calculate all the stats
			self.LeftBoutStats = StatBot(self.LeftBouts)
			self.LeftInterboutStats = StatBot(self.LeftInterbouts)
			self.RightBoutStats = StatBot(self.RightBouts)
			self.RightInterboutStats = StatBot(self.RightInterbouts)

		self.IsProcessed = True

	def PlotVsTime(self, fname, minutes=1):
		"""
		Plot bouts against time. Bouts are grouped by the minute.
		Set @minutes to something other than 1 to pre-group them into larger groups
		"""

		fig,axes = pyplot.subplots(2)
		fig.suptitle("VsTime for %s" % fname)

		axes[0].set_ylabel("Left (# Bouts)")
		axes[1].set_ylabel("Right (# Bouts)")
		axes[1].set_xlabel("Time (min)")
		fig.autofmt_xdate()

		if minutes != 1:
			raise NotImplementedError("Minutes group not implemented yet")

		start = self.Spandt[0]
		end = self.Spandt[1]

		# -------- LEFT --------
		# Get all the datetime objects and sort them
		keys = list(self.LeftVsTime.keys())
		keys.sort()

		# X data is datetime values
		# Y data is bout counts per time
		xl = []
		yl = []
		# Iterate over the whole time so that zeroes can be injected to plot nicely
		for minute in range(0, int((end-start).total_seconds()/60)+1):
			dt = start + datetime.timedelta(minutes=minute)
			xl.append(dt)
			if dt in keys:
				yl.append(len(self.LeftVsTime[dt]))
			else:
				# If there isn't a data point at this time, then there were zero bouts
				yl.append(0)


		# -------- RIGHT --------
		# Should mirror left except by the use of RightVsTime instead
		keys = list(self.RightVsTime.keys())
		keys.sort()

		xr = []
		yr = []
		for minute in range(0, int((end-start).total_seconds()/60)+1):
			dt = start + datetime.timedelta(minutes=minute)
			xr.append(dt)
			if dt in keys:
				yr.append(len(self.RightVsTime[dt]))
			else:
				yr.append(0)

		# Want y-axis on both to match
		maxy = max(max(yl), max(yr))
		axes[0].set_ylim(0,maxy)
		axes[1].set_ylim(0,maxy)

		axes[0].plot(xl, yl)
		axes[1].plot(xr, yr)

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotBoutRepetitions(self, fname, minutes=1):
		"""
		Plot a cumulative bout count for each tube that is reset when the other tube is used.
		"""

		fig,axes = pyplot.subplots(2)
		fig.suptitle("Bout Repititions for %s" % fname)

		axes[0].set_ylabel("Left (# Bouts)")
		axes[1].set_ylabel("Right (# Bouts)")
		axes[1].set_xlabel("Time (min)")
		fig.autofmt_xdate()

		if minutes != 1:
			raise NotImplementedError("Minutes group not implemented yet")

		start = self.Spandt[0]
		end = self.Spandt[1]

		# -------- LEFT vs RIGHT --------
		# Count number of left bouts that is reset when a right bout is encountered
		# Get all the datetime objects and sort them
		lkeys = list(self.LeftVsTime.keys())
		lkeys.sort()
		rkeys = list(self.RightVsTime.keys())
		rkeys.sort()

		# X data is datetime values
		# Y data is bout counts per time
		xl = []
		yl = []
		# Iterate over the whole time so that zeroes can be injected to plot nicely
		cnt = 0
		for minute in range(0, int((end-start).total_seconds()/60)+1):
			dt = start + datetime.timedelta(minutes=minute)
			xl.append(dt)
			if dt in rkeys:
				cnt = 0
				yl.append(cnt)

			elif dt in lkeys:
				cnt += 1
				yl.append(cnt)

			else:
				# If there isn't a data point at this time, then there were zero bouts
				yl.append(cnt)


		# -------- RIGHT vs LEFT --------
		# Count number of right bouts that is reset when a left bout is encountered
		# X data is datetime values
		# Y data is bout counts per time
		xr = []
		yr = []
		# Iterate over the whole time so that zeroes can be injected to plot nicely
		cnt = 0
		for minute in range(0, int((end-start).total_seconds()/60)+1):
			dt = start + datetime.timedelta(minutes=minute)
			xr.append(dt)
			if dt in lkeys:
				cnt = 0
				yr.append(cnt)

			elif dt in rkeys:
				cnt += 1
				yr.append(cnt)

			else:
				# If there isn't a data point at this time, then there were zero bouts
				yr.append(cnt)

		# Want y-axis on both to match
		maxy = max(max(yl), max(yr))
		axes[0].set_ylim(0,maxy)
		axes[1].set_ylim(0,maxy)

		axes[0].plot(xl, yl)
		axes[1].plot(xr, yr)

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotCumulativeBoutTimes(self, fname):
		"""
		Plot cumulative bout times.
		"""

		fig,axes = pyplot.subplots(1)
		fig.autofmt_xdate()
		fig.suptitle("Cumulative Bout Times for %s" % fname)

		axes.set_xlabel("Time (ms)")
		axes.set_ylabel("Cumulative Time (sec)")
		# Data is in milliseconds, so divide each point by 1000.0 to get seconds

		x = [_[0] for _ in self.LeftCumulative]
		y = [_[1]/1000.0 for _ in self.LeftCumulative]
		axes.plot(x,y, 'r', label="Left")

		x = [_[0] for _ in self.RightCumulative]
		y = [_[1]/1000.0 for _ in self.RightCumulative]
		axes.plot(x,y, 'b', label="Right")

		axes.legend(loc="lower right")

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotCumulativeNormalizedVolume(self, fname):
		"""
		Plot cumulative volume times normalized to recorded volume over the day.
		"""

		fig,axes = pyplot.subplots(1)
		fig.autofmt_xdate()
		fig.suptitle("Cumulative Normalized Volume for %s" % fname)

		axes.set_xlabel("Time (ms)")
		axes.set_ylabel("Cumulative Volume (mL)")

		x = [_[0] for _ in self.LeftCumulativeTotalVolume]
		y = [_[1] for _ in self.LeftCumulativeTotalVolume]
		axes.plot(x,y, 'r', label="Left")

		x = [_[0] for _ in self.RightCumulativeTotalVolume]
		y = [_[1] for _ in self.RightCumulativeTotalVolume]
		axes.plot(x,y, 'b', label="Right")

		axes.legend(loc="lower right")

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

		with open(fname + '-left.csv', 'w', newline='') as f:
			w = csv.writer(f)
			for dt,val in self.LeftCumulativeTotalVolume:
				w.writerow([dt,val])

		with open(fname + '-right.csv', 'w', newline='') as f:
			w = csv.writer(f)
			for dt,val in self.RightCumulativeTotalVolume:
				w.writerow([dt,val])

	def PlotBoutBoxplot(self, fname, limitextremes=True):
		"""
		Plot bouts as a box plot.
		@limitextremes, if True, then the ultra extreme outliers are chopped off by fixing the y-axis limits.
		"""

		fig,axes = pyplot.subplots(1)
		fig.autofmt_xdate()
		fig.suptitle("Box Plot for %s" % fname)

		axes.set_ylabel('BoutTime (ms)')
		axes.boxplot( [self.LeftBouts, self.RightBouts], labels=['Left', 'Right'])

		if limitextremes:
			# With extreme outliers, the box plot gets squished into nohting so this limits the y-axis
			# to the min/max data point or no more than the definition of "extreme outlier" of Q25-3*IQR or Q75+3*IQR
			# This will remove outliers from the plot bu that makes the plot more useful

			if self.LeftBoutStats.Length != 0 and self.RightBoutStats.Length != 0:
				ymin_l = max(self.LeftBoutStats.Minimum, self.LeftBoutStats.Quartile25 - 3*self.LeftBoutStats.IQR)
				ymin_r = max(self.RightBoutStats.Minimum, self.RightBoutStats.Quartile25 - 3*self.RightBoutStats.IQR)
				ymin = min(ymin_l, ymin_r)

				ymax_l = min(self.LeftBoutStats.Maximum, self.LeftBoutStats.Quartile75 + 3*self.LeftBoutStats.IQR)
				ymax_r = min(self.RightBoutStats.Maximum, self.RightBoutStats.Quartile75 + 3*self.RightBoutStats.IQR)
				ymax = max(ymax_l, ymax_r)

			elif self.LeftBoutStats.Length == 0:
				ymin = max(self.RightBoutStats.Minimum, self.RightBoutStats.Quartile25 - 3*self.RightBoutStats.IQR)
				ymax = min(self.RightBoutStats.Maximum, self.RightBoutStats.Quartile75 + 3*self.RightBoutStats.IQR)
			elif self.RightBoutStats.Length == 0:
				ymin = max(self.RightBoutStats.Minimum, self.RightBoutStats.Quartile25 - 3*self.RightBoutStats.IQR)
				ymax = min(self.RightBoutStats.Maximum, self.RightBoutStats.Quartile75 + 3*self.RightBoutStats.IQR)
			else:
				raise NotImplementedError("Plotting no data?")

			# Fudge the bounds a little just to add some padding
			ymin *= 0.95
			ymax *= 1.05

			axes.set_ylim((ymin,ymax))

		else:
			pass

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotBoutHistogram_Overlap(self, fname, bins=25):
		"""
		Plot bouts as a histogram with @bins of data.
		Left and right data plots are shown on the same axes (overlapping).
		"""

		fig,axes = pyplot.subplots(1)
		fig.autofmt_xdate()
		fig.suptitle("Bout Histogram for %s" % fname)

		colors = ['blue', 'orange']
		axes.hist( [self.LeftBouts, self.RightBouts], bins=bins, color=colors, label=['Left', 'Right'])
		axes.set_xlabel('Bins of Time (ms)')
		axes.legend(loc='upper right')

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotBoutHistogram_SideBySide(self, fname, bins=25):
		"""
		Plot bouts as a histogram with @bins of data.
		Left and right data plots are shown side-by-side as separate plots.
		"""

		fig,axes = pyplot.subplots(1,2)
		fig.autofmt_xdate()
		fig.suptitle("Bout Histogram for %s" % fname)

		axes[0].set_xlabel('Left (Bins of Time (ms))')
		axes[1].set_xlabel('Right (Bins of Time (ms))')
		axes[0].hist(self.LeftBouts, bins=bins, color='blue')
		axes[1].hist(self.RightBouts, bins=bins, color='orange')

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotInterboutHistogram_Overlap(self, fname, bins=25):
		"""
		Plot interbouts as a histogram with @bins of data.
		Left and right data plots are shown on the same axes (overlapping).
		"""

		fig,axes = pyplot.subplots(1)
		fig.autofmt_xdate()
		fig.suptitle("Interbout Histogram for %s" % fname)

		colors = ['blue', 'orange']
		axes.hist( [self.LeftInterbouts, self.RightInterbouts], bins=bins, color=colors, label=['Left', 'Right'])
		axes.set_xlabel('Bins of Time (ms)')
		axes.legend(loc='upper right')

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	def PlotInterboutHistogram_SideBySide(self, fname, bins=25):
		"""
		Plot interbouts as a histogram with @bins of data.
		Left and right data plots are shown side-by-side as separate plots.
		"""

		fig,axes = pyplot.subplots(1,2)
		fig.autofmt_xdate()
		fig.suptitle("Interbout Histogram for %s" % fname)

		axes[0].set_xlabel('Left (Bins of Time (ms))')
		axes[1].set_xlabel('Right (Bins of Time (ms))')
		axes[0].hist(self.LeftInterbouts, bins=bins, color='blue')
		axes[1].hist(self.RightInterbouts, bins=bins, color='orange')

		# SAVE IT
		fig.savefig(fname)
		pyplot.close()

	@staticmethod
	def PlotStatsTable(fname, *objs):
		wb = Workbook()
		ws = wb.active
		ws.title = "Stats"

		# Section labels
		ws['A1'] = 'Bouts'
		ws['A23'] = 'Interbouts'

		# Same row headers for bouts and interbouts
		header = ['N', 'Sum', 'Min', 'Max', 'Mean', 'Q25', 'Q50', 'Q75', 'IQR']

		# Put in device IDs as columns for each section
		for pos in ([2,5], [24,5]): # E2 and E24
			for obj in objs:
				ws.cell(pos[0],pos[1]+0).value = obj.DeviceID
				pos[1] += 1

		# Make headers for each section
		for pos in ([3,2], [25,2]): #B3 and B25
			for h in header:
				# Header info
				ws.cell(pos[0]+0, pos[1]).value = h
				ws.cell(pos[0]+0, pos[1]+1).value = 'Left'
				ws.cell(pos[0]+1, pos[1]+1).value = 'Right'

				col_a = get_column_letter(pos[1]+2+1)
				col_b = get_column_letter(pos[1]+2+1 + len(objs)-1)

				#ws.cell(pos[0]+0, pos[1]+2).value = '=ttest(%s%d:%s%d, %s%d:%s%d, 2,1)' % (col_a,pos[0]+0, col_b,pos[0]+0, col_a,pos[0]+1, col_b,pos[0]+1)

				pos[0] += 2

		# Inject data
		for cnt,obj in enumerate(objs):
			# Increment row for each object
			pos = [3,5+cnt] # E3

			idx = 0

			left = obj.LeftBoutStats
			right = obj.RightBoutStats

			attrs = ['Length', 'Sum', 'Minimum', 'Maximum', 'Mean', 'Quartile25', 'Median', 'Quartile75', 'IQR']
			possnull = ['Minimum', 'Maximum', 'Mean', 'Quartile25', 'Median', 'Quartile75', 'IQR']
			for attr in attrs:
				l = getattr(left, attr)
				r = getattr(right, attr)

				if attr in possnull:
					ws.cell(pos[0]+idx+0, pos[1]).value = l or ""
					ws.cell(pos[0]+idx+1, pos[1]).value = r or ""
				else:
					ws.cell(pos[0]+idx+0, pos[1]).value = l
					ws.cell(pos[0]+idx+1, pos[1]).value = r
				idx += 2

		# Inject data
		for cnt,obj in enumerate(objs):
			# Increment row for each object
			pos = [25,5+cnt] # E3

			idx = 0

			left = obj.LeftInterboutStats
			right = obj.RightInterboutStats

			attrs = ['Length', 'Sum', 'Minimum', 'Maximum', 'Mean', 'Quartile25', 'Median', 'Quartile75', 'IQR']
			possnull = ['Minimum', 'Maximum', 'Mean', 'Quartile25', 'Median', 'Quartile75', 'IQR']
			for attr in attrs:
				l = getattr(left, attr)
				r = getattr(right, attr)

				if attr in possnull:
					ws.cell(pos[0]+idx+0, pos[1]).value = l or ""
					ws.cell(pos[0]+idx+1, pos[1]).value = r or ""
				else:
					ws.cell(pos[0]+idx+0, pos[1]).value = l
					ws.cell(pos[0]+idx+1, pos[1]).value = r
				idx += 2

		try:
			os.unlink(fname)
		except:
			pass
		wb.save(fname)
